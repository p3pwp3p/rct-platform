# -*- coding: utf-8 -*-
"""
RCT 카피자 동기화 (단일 프로그램)

- Playwright 없이 Chrome 디버그 포트(CDP)에 직접 연결 → 가벼운 단일 exe.
- GUI 모드(기본): 창을 띄워 버튼으로 실행.
- 자동 모드(--auto): 창 없이 1회 실행(작업 스케줄러용).

설정은 exe(또는 이 파일) 옆의 config.ini 에서 읽는다.
"""
import os
import sys
import json
import time
import threading
import subprocess
import configparser
import urllib.request
import urllib.error

try:
    from websocket import create_connection  # websocket-client
except Exception:
    create_connection = None

APP_NAME = "RCT 카피자 동기화"

# ── 경로/설정 ────────────────────────────────────────────────────────────────
def base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def export_dir():
    d = os.path.join(base_dir(), "exports")
    os.makedirs(d, exist_ok=True)
    return d

def list_exports():
    """exports 폴더의 엑셀 파일을 최신순으로 (파일명, 경로, 수정시각, 크기KB) 반환."""
    d = export_dir()
    items = []
    for name in os.listdir(d):
        if not name.lower().endswith(".xlsx"):
            continue
        p = os.path.join(d, name)
        try:
            st = os.stat(p)
            items.append((name, p, st.st_mtime, round(st.st_size / 1024, 1)))
        except OSError:
            pass
    items.sort(key=lambda x: x[2], reverse=True)
    return items

def load_config():
    cfg = configparser.ConfigParser()
    path = os.path.join(base_dir(), "config.ini")
    defaults = {
        "api_url": "http://localhost:3000",
        "internal_secret": "",
        "strategy_url": "https://secure.vantagemarkets.com/copyTrading/discover/discoverDetail?strategyId=1289377&tab=2",
        "login_url": "https://secure.vantagemarkets.com/login",
        "apply": "false",
        "debug_port": "9222",
        "chrome_path": r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        "profile_dir": ".chrome-profile",
    }
    if os.path.exists(path):
        cfg.read(path, encoding="utf-8-sig")   # BOM 있어도 안전
        if cfg.has_section("rct"):
            for k in defaults:
                if cfg.has_option("rct", k):
                    defaults[k] = cfg.get("rct", k)
    return defaults

# ── 추출 JS (카피자 표) ──────────────────────────────────────────────────────
EXTRACT_JS = r"""
(() => {
  const text = document.body.innerText;
  const rows = [];
  const lines = text.split('\n').map(s => s.trim()).filter(Boolean);
  for (let i = 0; i < lines.length; i++) {
    const idm = lines[i].match(/ID:\s*(\d{5,})(?:\((\d+)\))?/);
    if (!idm) continue;
    const nick = (lines[i-1] || '').replace(/\s+\d{5,}$/, '');
    const win = lines.slice(i, i + 8).join(' | ');
    const usd = [...win.matchAll(/(-?[\d,]+\.\d{2})\s*USD/g)].map(m => parseFloat(m[1].replace(/,/g,'')));
    rows.push({
      nickname: nick, masterAccountNo: idm[1], subAccountNo: idm[2] || null,
      feeBalance: usd.length>0?usd[0]:null, realizedProfit: usd.length>1?usd[1]:null,
      group: (win.match(/권하다\s*\d+/)||[null])[0],
    });
  }
  return rows;
})()
"""
READY_JS = "(() => /ID:\\s*\\d{5,}/.test(document.body.innerText))()"

MIN_PER_NODE = 3000

def save_excel(rows, path, log):
    """추출 결과를 엑셀(.xlsx)로 저장. openpyxl 필요."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        log("openpyxl 미설치로 엑셀 저장 건너뜀(copiers.json 은 저장됨).")
        return None
    wb = Workbook(); ws = wb.active; ws.title = "카피자"
    headers = ["닉네임", "Vantage C.T", "서브계정", "수수료잔고(USD)", "실현수익(USD)", "허용노드(잔고/3000)", "그룹"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="11141B")
    for c in ws[1]:
        c.font = Font(bold=True, color="4DB6AC"); c.fill = head_fill; c.alignment = Alignment(horizontal="center")
    for r in rows:
        bal = r.get("feeBalance")
        allowed = int(bal // MIN_PER_NODE) if isinstance(bal, (int, float)) else None
        ws.append([r.get("nickname"), r.get("masterAccountNo"), r.get("subAccountNo"),
                   bal, r.get("realizedProfit"), allowed, r.get("group")])
    # 열 너비
    widths = [16, 14, 12, 16, 15, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return path

# ── CDP 클라이언트 ───────────────────────────────────────────────────────────
class CDP:
    def __init__(self, ws_url, timeout=30):
        # 최신 Chrome 은 Origin 헤더가 있으면 ws 연결을 거부(403) → Origin 을 안 보냄
        self.ws = create_connection(ws_url, max_size=None, suppress_origin=True)
        self.ws.settimeout(timeout)
        self._id = 0
    def cmd(self, method, params=None, timeout=30):
        self._id += 1
        mid = self._id
        self.ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
        deadline = time.time() + timeout
        while time.time() < deadline:
            msg = json.loads(self.ws.recv())
            if msg.get("id") == mid:
                if "error" in msg:
                    raise RuntimeError(msg["error"].get("message", str(msg["error"])))
                return msg.get("result", {})
        raise TimeoutError(method)
    def evaluate(self, expr, timeout=30):
        r = self.cmd("Runtime.evaluate", {"expression": expr, "returnByValue": True, "awaitPromise": True}, timeout)
        res = r.get("result", {})
        if res.get("subtype") == "error":
            raise RuntimeError(res.get("description", "JS error"))
        return res.get("value")
    def close(self):
        try: self.ws.close()
        except Exception: pass

# ── Chrome 기동/탐지 ─────────────────────────────────────────────────────────
def debug_alive(port):
    try:
        with urllib.request.urlopen(f"http://localhost:{port}/json/version", timeout=2):
            return True
    except Exception:
        return False

def kill_profile_chrome(profile, log):
    """우리 디버그 프로필을 물고 있는 잔여 chrome 만 종료(일반 Chrome 은 안 건드림)."""
    token = os.path.basename(profile.rstrip("\\/"))
    if not token:
        return
    ps = ("Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
          f"Where-Object {{ $_.CommandLine -like '*{token}*' }} | "
          "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }")
    try:
        flags = 0x08000000  # CREATE_NO_WINDOW (콘솔 깜빡임 방지)
        subprocess.run(["powershell", "-NoProfile", "-Command", ps], creationflags=flags, timeout=15)
        log("잔여 디버그 Chrome 정리 완료.")
        time.sleep(2)
    except Exception as e:
        log(f"(프로필 정리 스킵: {e})")

def launch_chrome(cfg, log):
    chrome = cfg["chrome_path"]
    if not os.path.exists(chrome):
        alt = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        chrome = alt if os.path.exists(alt) else chrome
    if not os.path.exists(chrome):
        raise FileNotFoundError("Chrome 실행 파일을 찾을 수 없습니다. config.ini 의 chrome_path 를 확인하세요.")
    profile = cfg["profile_dir"]
    if not os.path.isabs(profile):
        profile = os.path.join(base_dir(), profile)
    # 포트가 죽어있는데 프로필을 물고 있는 잔여 인스턴스 → 먼저 정리
    kill_profile_chrome(profile, log)
    log(f"Chrome 실행(포트 {cfg['debug_port']})…")
    subprocess.Popen([chrome, f"--remote-debugging-port={cfg['debug_port']}",
                      "--remote-allow-origins=*",
                      f"--user-data-dir={profile}", cfg["login_url"]])
    for _ in range(20):
        if debug_alive(cfg["debug_port"]):
            return
        time.sleep(1)
    raise RuntimeError("Chrome 디버그 포트가 안 열렸습니다. 열려있는 모든 Chrome 창(작업관리자 chrome.exe 포함)을 닫고 다시 시도하세요.")

def find_page_ws(port):
    with urllib.request.urlopen(f"http://localhost:{port}/json", timeout=5) as r:
        targets = json.loads(r.read().decode("utf-8"))
    pages = [t for t in targets if t.get("type") == "page" and t.get("webSocketDebuggerUrl")]
    for t in pages:
        if "vantagemarkets.com" in (t.get("url") or ""):
            return t["webSocketDebuggerUrl"]
    return pages[0]["webSocketDebuggerUrl"] if pages else None

# ── 서버 전송 ────────────────────────────────────────────────────────────────
def send_to_server(cfg, rows, log):
    api, secret = cfg["api_url"].rstrip("/"), cfg["internal_secret"]
    if not api or not secret:
        log("서버 전송 건너뜀(api_url/internal_secret 미설정).")
        return None
    copiers = [{"vantageCt": r["masterAccountNo"], "feeBalance": r.get("feeBalance"), "realizedProfit": r.get("realizedProfit")}
               for r in rows if r.get("feeBalance") is not None]
    apply = str(cfg.get("apply", "false")).lower() == "true"
    payload = json.dumps({"copiers": copiers, "apply": apply, "graceByDeposit": True}).encode("utf-8")
    req = urllib.request.Request(f"{api}/api/admin/import-copiers", data=payload, method="POST",
                                 headers={"Content-Type": "application/json", "x-internal-secret": secret})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        log(f"전송 실패(HTTP {e.code}): {e.read().decode('utf-8','ignore')[:200]}")
        return None
    except Exception as e:
        log(f"전송 실패: {e}")
        return None
    s = body.get("summary", {})
    log(f"대사: 매칭 {s.get('matched')} / 미매칭 {s.get('unmatched')} / 예고시작 {s.get('graceStart')} / "
        f"정지 {s.get('toSuspend')} / 재활성 {s.get('toReactivate')}  (적용={s.get('applied')})")
    return body

# ── 동기화 1회 ───────────────────────────────────────────────────────────────
def run_sync(cfg, log, allow_launch=True):
    if create_connection is None:
        raise RuntimeError("websocket-client 미설치. pip install websocket-client")
    port = cfg["debug_port"]
    if not debug_alive(port):
        if not allow_launch:
            raise RuntimeError("Chrome 이 열려있지 않습니다. 먼저 [Chrome 열기] 를 눌러 로그인한 뒤 다시 실행하세요.")
        launch_chrome(cfg, log)
        log("※ 처음이면 열린 Chrome 에서 Vantage 로그인 후 다시 실행하세요.")
    ws = find_page_ws(port)
    if not ws:
        raise RuntimeError("Chrome 페이지 탭을 찾지 못했습니다.")
    cdp = CDP(ws)
    try:
        log("카피자 탭으로 이동…")
        cdp.cmd("Page.navigate", {"url": cfg["strategy_url"]})
        # 표 로딩 대기(최대 30초)
        ready = False
        for _ in range(30):
            time.sleep(1)
            try:
                if cdp.evaluate(READY_JS):
                    ready = True; break
            except Exception:
                pass
        if not ready:
            log("표를 찾지 못함(로그인 안 됐거나 화면이 다름).")
            return None
        rows = cdp.evaluate(EXTRACT_JS) or []
        log(f"추출된 카피자: {len(rows)}명")
        # 파일 저장 (JSON 최신본 + 날짜별 엑셀 누적)
        out = os.path.join(base_dir(), "copiers.json")
        with open(out, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        exdir = export_dir()
        stamp = time.strftime("%Y-%m-%d_%H%M")
        xlsx = save_excel(rows, os.path.join(exdir, f"카피자_{stamp}.xlsx"), log)
        if xlsx:
            log(f"엑셀 저장: exports/{os.path.basename(xlsx)}")
        server = send_to_server(cfg, rows, log) if rows else None
        log("완료.")
        return {"rows": rows, "server": server}
    finally:
        cdp.close()

# ── GUI ──────────────────────────────────────────────────────────────────────
def run_gui(cfg):
    import tkinter as tk
    from tkinter import scrolledtext, messagebox, ttk

    root = tk.Tk()
    root.title(APP_NAME)
    root.geometry("620x520")
    root.minsize(520, 420)
    root.configure(bg="#11141b")

    last = {"result": None, "t": 0.0}   # 마지막 실행 결과/시각

    # 상단(제목 + 체크박스)
    top = tk.Frame(root, bg="#11141b"); top.pack(side="top", fill="x", padx=14, pady=(14, 4))
    tk.Label(top, text=APP_NAME, fg="#e0e6ed", bg="#11141b", font=("Malgun Gothic", 13, "bold")).pack(side="left")
    apply_var = tk.BooleanVar(value=str(cfg.get("apply", "false")).lower() == "true")
    tk.Checkbutton(top, text="실제 적용(정지/재활성)", variable=apply_var, fg="#94a3b8", bg="#11141b",
                   selectcolor="#11141b", activebackground="#11141b", activeforeground="#e0e6ed").pack(side="right")

    # 요약 패널
    summ = tk.Frame(root, bg="#0a0c10"); summ.pack(side="top", fill="x", padx=14, pady=(2, 6))
    summary_var = tk.StringVar(value="아직 실행 안 함 — [동기화 실행]을 누르세요.")
    tk.Label(summ, textvariable=summary_var, fg="#e0e6ed", bg="#0a0c10", anchor="w",
             font=("Malgun Gothic", 10), padx=12, pady=10).pack(fill="x")

    # 하단(버튼 + 상태) — 로그보다 먼저 배치해 항상 보이도록 고정
    btns = tk.Frame(root, bg="#11141b"); btns.pack(side="bottom", fill="x", padx=14, pady=(4, 14))
    run_btn = tk.Button(btns, text="동기화 실행", bg="#4db6ac", fg="#07080a",
                        font=("Malgun Gothic", 10, "bold"), relief="flat", padx=18, pady=8)
    run_btn.pack(side="right")
    chrome_btn = tk.Button(btns, text="Chrome 열기", bg="#242a35", fg="#e0e6ed",
                           font=("Malgun Gothic", 10), relief="flat", padx=14, pady=8)
    chrome_btn.pack(side="right", padx=(0, 8))
    # 자동 반복
    auto_var = tk.BooleanVar(value=False)
    interval_var = tk.StringVar(value="60")
    autofr = tk.Frame(btns, bg="#11141b"); autofr.pack(side="right", padx=(0, 12))
    tk.Checkbutton(autofr, text="자동", variable=auto_var, fg="#94a3b8", bg="#11141b",
                   selectcolor="#11141b", activebackground="#11141b", activeforeground="#e0e6ed").pack(side="left")
    tk.Entry(autofr, textvariable=interval_var, width=3, bg="#0a0c10", fg="#e0e6ed",
             insertbackground="#e0e6ed", relief="flat", justify="center").pack(side="left")
    tk.Label(autofr, text="분", fg="#64748b", bg="#11141b").pack(side="left")

    list_btn = tk.Button(btns, text="📁 엑셀 목록", bg="#242a35", fg="#e0e6ed",
                         font=("Malgun Gothic", 9), relief="flat", padx=10, pady=6)
    list_btn.pack(side="left")
    recon_btn = tk.Button(btns, text="📋 대사결과", bg="#242a35", fg="#e0e6ed",
                          font=("Malgun Gothic", 9), relief="flat", padx=10, pady=6)
    recon_btn.pack(side="left", padx=(8, 0))
    status = tk.Label(btns, text="대기 중", fg="#64748b", bg="#11141b", font=("Malgun Gothic", 9))
    status.pack(side="left", padx=(10, 0))

    # 로그(가운데, 남는 공간 채움)
    logbox = scrolledtext.ScrolledText(root, bg="#0a0c10", fg="#94a3b8", insertbackground="#94a3b8",
                                       font=("Consolas", 9), relief="flat")
    logbox.pack(side="top", fill="both", expand=True, padx=14, pady=6)

    def log(msg):
        logbox.insert("end", msg + "\n"); logbox.see("end"); root.update_idletasks()

    def update_summary():
        r = last["result"]
        if not r:
            return
        rows = r.get("rows") or []
        s = (r.get("server") or {}).get("summary") or {}
        pend = (s.get("graceStart", 0) or 0) + (s.get("toSuspend", 0) or 0)
        unm = s.get("unmatched", 0) or 0
        applied = "적용됨" if s.get("applied") else "시뮬레이션"
        t = time.strftime("%H:%M", time.localtime(last["t"]))
        summary_var.set(f"카피자 {len(rows)}명  ·  정지예정 {pend}  ·  미매칭 {unm}  ·  {applied}  ·  마지막 {t}")

    busy = {"v": False}
    def do_run(auto=False):
        if busy["v"]:
            return
        if apply_var.get() and not auto:
            if not messagebox.askyesno("실제 적용 확인",
                    "‘실제 적용’이 켜져 있습니다.\n증거금 미달 회원 노드가 실제로 정지 예고/정지될 수 있습니다.\n\n계속할까요?"):
                return
        busy["v"] = True; status.config(text="동기화 중…", fg="#4db6ac")
        run_cfg = dict(cfg); run_cfg["apply"] = "true" if apply_var.get() else "false"
        def work():
            try:
                res = run_sync(run_cfg, log, allow_launch=False)
                last["result"] = res; last["t"] = time.time()
                update_summary()
                status.config(text="완료", fg="#34d399")
            except Exception as e:
                log(f"오류: {e}"); status.config(text="오류", fg="#f87171")
            finally:
                busy["v"] = False
        threading.Thread(target=work, daemon=True).start()

    def on_auto_toggle():
        if auto_var.get() and apply_var.get():
            if not messagebox.askyesno("자동 + 실제 적용",
                    "자동 반복이 ‘실제 적용’과 함께 켜집니다.\n매 실행마다 회원 노드가 자동으로 정지 처리될 수 있습니다.\n\n계속할까요?"):
                auto_var.set(False); return
        if auto_var.get():
            log(f"자동 반복 ON ({interval_var.get()}분마다)")
        else:
            log("자동 반복 OFF")

    def auto_tick():
        if auto_var.get() and not busy["v"]:
            try: mins = max(1, int(interval_var.get()))
            except Exception: mins = 60
            if time.time() - last["t"] >= mins * 60:
                do_run(auto=True)
        root.after(20000, auto_tick)   # 20초마다 조건 확인

    def show_recon():
        r = last["result"]
        win = tk.Toplevel(root); win.title("대사결과"); win.geometry("560x460"); win.configure(bg="#11141b")
        tk.Label(win, text="동기화 대사결과 (계정별 조치)", fg="#e0e6ed", bg="#11141b",
                 font=("Malgun Gothic", 11, "bold")).pack(anchor="w", padx=14, pady=(12, 6))
        recon = ((r or {}).get("server") or {}).get("reconciliation") or []
        if not recon:
            tk.Label(win, text="대사결과가 없습니다. 동기화를 먼저 실행하세요.",
                     fg="#64748b", bg="#11141b", font=("Malgun Gothic", 10)).pack(padx=14, pady=20)
            return
        cols = ("상태", "이름", "CT", "잔고", "허용", "활성")
        tv = ttk.Treeview(win, columns=cols, show="headings", height=16)
        for c, w in zip(cols, (80, 110, 90, 80, 50, 50)):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        label = {"grace_start": "정지예정", "suspend": "정지", "grace_cancel": "예고해제",
                 "reactivate": "재활성", "none": "정상", "unmatched": "미매칭"}
        order = {"suspend": 0, "grace_start": 1, "unmatched": 2, "grace_cancel": 3, "reactivate": 4, "none": 5}
        for row in sorted(recon, key=lambda x: order.get(x.get("action"), 9)):
            tv.insert("", "end", values=(
                label.get(row.get("action"), row.get("action")),
                row.get("name") or "-", row.get("vantageCt") or "-",
                f"${row.get('feeBalance'):,.0f}" if isinstance(row.get("feeBalance"), (int, float)) else "-",
                "∞" if row.get("allowedNodes") is None and row.get("action") != "unmatched" else (row.get("allowedNodes") if row.get("allowedNodes") is not None else "-"),
                row.get("activeNodes", "-"),
            ))
        tv.pack(fill="both", expand=True, padx=14, pady=6)

    def do_chrome():
        if busy["v"]:
            return
        busy["v"] = True; status.config(text="Chrome 여는 중…", fg="#4db6ac")
        def work():
            try:
                if debug_alive(cfg["debug_port"]):
                    log("Chrome 이 이미 열려있습니다(디버그 포트 활성).")
                else:
                    launch_chrome(cfg, log)
                    log("Chrome 을 열었습니다. Vantage 에 로그인한 뒤 '동기화 실행'을 누르세요.")
                status.config(text="대기 중", fg="#64748b")
            except Exception as e:
                log(f"오류: {e}"); status.config(text="오류", fg="#f87171")
            finally:
                busy["v"] = False
        threading.Thread(target=work, daemon=True).start()

    def show_exports():
        win = tk.Toplevel(root); win.title("엑셀 목록 (날짜별)"); win.geometry("460x420"); win.configure(bg="#11141b")
        tk.Label(win, text="저장된 카피자 엑셀 (최신순)", fg="#e0e6ed", bg="#11141b",
                 font=("Malgun Gothic", 11, "bold")).pack(anchor="w", padx=14, pady=(12, 6))
        lb = tk.Listbox(win, bg="#0a0c10", fg="#94a3b8", font=("Consolas", 9), relief="flat",
                        selectbackground="#4db6ac", selectforeground="#07080a", activestyle="none")
        lb.pack(fill="both", expand=True, padx=14, pady=6)
        files = list_exports()
        cur_date = None
        rowmap = {}  # listbox index → 파일 경로
        for name, path, mtime, kb in files:
            d = time.strftime("%Y-%m-%d", time.localtime(mtime))
            if d != cur_date:
                lb.insert("end", f"── {d} ──"); lb.itemconfig("end", fg="#64748b"); cur_date = d
            t = time.strftime("%H:%M", time.localtime(mtime))
            idx = lb.size()
            lb.insert("end", f"   {t}   {name}   ({kb} KB)")
            rowmap[idx] = path
        if not files:
            lb.insert("end", "저장된 엑셀이 없습니다. 동기화를 먼저 실행하세요.")

        def open_selected(_=None):
            sel = lb.curselection()
            if sel and sel[0] in rowmap:
                try: os.startfile(rowmap[sel[0]])
                except Exception as e: log(f"열기 실패: {e}")
        lb.bind("<Double-Button-1>", open_selected)

        bar = tk.Frame(win, bg="#11141b"); bar.pack(fill="x", padx=14, pady=(4, 14))
        tk.Button(bar, text="폴더 열기", bg="#242a35", fg="#e0e6ed", font=("Malgun Gothic", 9),
                  relief="flat", padx=12, pady=6, command=lambda: os.startfile(export_dir())).pack(side="left")
        tk.Button(bar, text="선택 파일 열기", bg="#4db6ac", fg="#07080a", font=("Malgun Gothic", 9, "bold"),
                  relief="flat", padx=12, pady=6, command=open_selected).pack(side="right")

    run_btn.config(command=do_run)
    chrome_btn.config(command=do_chrome)
    list_btn.config(command=show_exports)
    recon_btn.config(command=show_recon)
    for w in autofr.winfo_children():
        if isinstance(w, tk.Checkbutton):
            w.config(command=on_auto_toggle)

    log("준비됨. 처음이면 [Chrome 열기] → 로그인 → [동기화 실행] 순서로 하세요.")
    log(f"서버: {cfg['api_url']}  · 적용기본: {cfg.get('apply')}")
    root.after(20000, auto_tick)
    root.mainloop()

# ── 진입점 ───────────────────────────────────────────────────────────────────
def main():
    cfg = load_config()
    if "--auto" in sys.argv:
        # 자동 모드: 창 없이 1회 실행 + 로그를 파일에 기록(스케줄러용)
        logpath = os.path.join(base_dir(), "sync_log.txt")
        def log(m):
            line = f"{time.strftime('%Y-%m-%d %H:%M:%S')} {m}"
            print("[sync]", m, flush=True)
            try:
                with open(logpath, "a", encoding="utf-8") as f:
                    f.write(line + "\n")
            except Exception:
                pass
        try:
            run_sync(cfg, log)
        except Exception as e:
            log(f"오류: {e}"); sys.exit(1)
    else:
        run_gui(cfg)

if __name__ == "__main__":
    main()
